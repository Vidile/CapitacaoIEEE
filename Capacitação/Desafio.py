import math
from openpyxl import load_workbook


def main():
    entradas = ['dummy']

    while len(entradas) < 4:

        # Openpyxl
        wb = load_workbook('exports-data.xlsx')
        ws = wb['Planilha1']

        # Entrada de dados do usuario
        Entrada = input(str('Insira um País: '))
        entradas.append(Entrada)

        # Arrays para os valores de importação e anos
        valores = []
        anos = []

        # Laço que me permite ler os dados da tabela e os colocar nos arrays
        for i in range(1, 1716):
            aux = 'E'+str(i)
            auxB = 'C'+str(i)
            auxC = 'D'+str(i)
            a = str(ws[aux].value)

            # comparador usando o Openpyxl
            if Entrada == a:
                valores.append(int(ws[auxB].value))
                anos.append(int(ws[auxC].value))

        # Pega o valor mais alto de importação dado o ano e seu indice
        B = max(valores)
        c = valores.index(B)

        # Pega a variavel com indice correspondente ao maior valor
        ano_max = anos[c]

        # Operações matematicas pedidas soma e média
        soma_import = sum(valores)
        media = soma_import/len(valores)

        # Saída de dados
        print('\n Nome do País:', Entrada,
              "\n Soma da importações do País:", soma_import, "\n Média das importações Anuais: ", media)


main()
